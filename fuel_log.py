import openpyxl # type: ignore
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from datetime import datetime, timedelta
import pandas as pd # type: ignore

# Define constants
START_DATE = datetime(2024, 8, 1)
START_MONTH=8
END_DATE = datetime(2025, 3, 31)
INITIAL_ODOMETER = 17569
INR_PER_KM = 10
WORK_RELATED_KM = 110
TRIP_PURPOSE = "Official"
CLIENT_NAME = "Blink Charging"
IS_WORK_TRAVEL = "Y"
PERSONAL_TRAVEL = ""

# File paths
uploaded_file_path = "Log_Book_FY_2024-25_template.xlsx"
output_file_path = "Financial_Year_2024_25_Log_Book.xlsx"

# Create a new workbook
workbook = openpyxl.Workbook()
workbook.remove(workbook.active)  # Remove the default sheet


# Function to create a sheet for a given month
def create_month_sheet(ws, year, month, odometer_start):
    # Generate the date range for the month
    first_day = datetime(year, month, 1)
    last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    date_range = pd.date_range(start=first_day, end=last_day)

    # Add headers
    headers = [
        ("A1", "Blink Charging Software Solutions India Private Limited"),
        # Basic Details
        ("A3", "BASIC DETAILS"),
        ("A5", "Name:"),
        ("B5", "Ashish Kumar"),
        ("A6", "Employee ID:"),
        ("B6", "BLINKIN065"),
        ("A7", "Department:"),
        ("B7", "Technology"),
        ("A8", "Manager:"),
        ("B8", "Ajay Singh"),
        # Vehicle Details
        ("D3", "VEHICLE DETAILS"),
        ("D4", "Make:"),
        ("E4", "Hyundai"),
        ("D5", "Model:"),
        ("E5", "Xcent"),
        ("D6", "Year:"),
        ("E6", "2018"),
        ("D7", "Registration:"),
        ("E7", "Delhi"),
        ("D8", "Engine Size:"),
        ("E8", "1199 CC"),
        # Odometer Reading
        ("G3", "ODOMETER READING"),
        ("G6", "Finanical year:"),
        ("H6", "2024-25"),
        ("G7", ""),
        ("G8", ""),
        ("H7", ""),
        ("H8", ""),
        # Empty Columns
        ("K3", ""),
        ("K4", ""),
        ("K5", ""),
        ("K6", ""),
        ("K7", ""),
        ("K8", ""),
        ("K9", ""),
        # Data Headers
        ("A10", "Date of Trip (DD/MM/YY)"),
        ("A11", "Start"),
        ("B11", "End"),
        ("C10", "Odometer Reading"),
        ("C11", "Start"),
        ("D11", "End"),
        ("E10", "Purpose of Trip"),
        ("F10", "Name of Client"),
        ("G10", "Work-related travel? (Y/N)"),
        ("H10", "Work-related Travel (KM)"),
        ("I10", "Personal Travel (KM)"),
        ("J10", "INR Per KM"),
        ("K10", "Amount (INR)"),
    ]

    # Set Font style
    font_style_1 = Font(
        size=18,  # Font size
        bold=True,  # Bold text
        name="Algerian"
    )
    font_style_2 = Font(
        size=12,  # Font size
        bold=True,  # Bold text
        underline="single"
    )
    font_style_3 = Font(
        size=11,  # Font size
        bold=True,  # Bold text
    )

    # Set background color
    fill_1 = PatternFill(
        fill_type="solid",  # Fill type (e.g., solid, gradient)
        start_color="B4C7E7",  # Background color (Blue)
        end_color="B4C7E7"  # End color (same for solid fill)
    )

    # Set borders
    thin_border_1 = Border(
        left=Side(border_style="thin", color="000000"),  # Left border
        right=Side(border_style="thin", color="000000"),  # Right border
        top=Side(border_style="thin", color="000000"),  # Top border
        bottom=Side(border_style="thin", color="000000")  # Bottom border
    )
    thin_border_2 = Border(
        right=Side(border_style="thin", color="000000"),  # Right border
    )

    # Set alignment properties
    alignment_1 = Alignment(
        horizontal="center",  # Center text horizontally
        vertical="center",  # Align text to the middle vertically
        wrap_text=True  # Enable text wrapping
    )
    alignment_2 = Alignment(
        horizontal="right",  # Center text horizontally
        vertical="center",  # Align text to the middle vertically
        wrap_text=True  # Enable text wrapping
    )
    alignment_3 = Alignment(
        horizontal="left",  # Center text horizontally
        vertical="center",  # Align text to the middle vertically
        wrap_text=True  # Enable text wrapping
    )

    for cell, text in headers:
        ws[cell] = text
        if cell == "A1":
            ws[cell].font = font_style_1
            ws[cell].fill = fill_1
            ws[cell].border = thin_border_1
            ws[cell].alignment = alignment_1
        elif cell in ("A3", "D3", "G3"):
            ws[cell].font = font_style_2
            ws[cell].alignment = alignment_1
        elif cell in ("A5", "A6", "A7", "A8", "D4", "D5", "D6", "D7", "D8", "G6", "G7", "G8"):
            ws[cell].font = font_style_3
            ws[cell].alignment = alignment_2
        elif cell in ("B5", "B6", "B7", "B8", "E4", "E5", "E6", "E7", "E8", "H6", "H7", "H8"):
            ws[cell].alignment = alignment_3
            ws[cell].border = thin_border_1
        elif cell in ("K3", "K4", "K5", "K6", "K7", "K8", "K9"):
            ws[cell].border = thin_border_2
        else:
            ws[cell].font = font_style_3
            ws[cell].fill = fill_1
            ws[cell].border = thin_border_1
            ws[cell].alignment = alignment_1

    # Merge cells
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=11)   # Top header Company name
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)    # Basic Details
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=5)    # Vehicle Details
    ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=8)    # Odometer Reading

    # Data headers
    ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=2)      # Date of Trip
    ws.merge_cells(start_row=10, start_column=3, end_row=10, end_column=4)      # Odometer Reading
    ws.merge_cells(start_row=10, start_column=5, end_row=11, end_column=5)      # Purpose of Trip
    ws.merge_cells(start_row=10, start_column=6, end_row=11, end_column=6)      # Name of Client
    ws.merge_cells(start_row=10, start_column=7, end_row=11, end_column=7)      # Work-related travel? (Y/N)
    ws.merge_cells(start_row=10, start_column=8, end_row=11, end_column=8)      # Work-related Travel (KM)
    ws.merge_cells(start_row=10, start_column=9, end_row=11, end_column=9)      # Personal Travel (KM)
    ws.merge_cells(start_row=10, start_column=10, end_row=11, end_column=10)    # INR Per KM
    ws.merge_cells(start_row=10, start_column=11, end_row=11, end_column=11)    # Amount (INR)

    # Add data to odometer fields above headers
    month_name = first_day.strftime("%B")
    last_date = last_day.strftime("%d")
    if last_date == "31":
        last_date += "st"
    else:
        last_date += "th"
    ws["G7"] = f"As at 1st of {month_name} {year}"
    ws["G8"] = f"As at {last_date} of {month_name} {year}"

    # Set odometer reading as on 1st of every month
    ws["H7"] = odometer_start

    # Fill in the data
    current_odometer = odometer_start
    total_cost = 0
    weekend_count = 0
    last_odometer_value = 0
    for i, date in enumerate(date_range, start=12):  # Start filling rows from the 12th row
        # Format the date
        date_str = date.strftime("%d/%m/%y")

        # Add dates to the columns
        ws[f"A{i}"] = date_str
        ws[f"B{i}"] = date_str

        # Add odometer readings
        ws[f"C{i}"] = current_odometer
        current_odometer += WORK_RELATED_KM
        ws[f"D{i}"] = current_odometer

        # Add work-related travel details
        ws[f"H{i}"] = WORK_RELATED_KM
        ws[f"J{i}"] = INR_PER_KM
        ws[f"K{i}"] = WORK_RELATED_KM * INR_PER_KM

        # Add data to other columns
        ws[f"E{i}"] = TRIP_PURPOSE
        ws[f"F{i}"] = CLIENT_NAME
        ws[f"G{i}"] = IS_WORK_TRAVEL
        ws[f"I{i}"] = PERSONAL_TRAVEL

        # saving the last odometer value of the month
        last_odometer_value = ws[f"D{i}"].value
        # Highlight weekends in red
        if date.weekday() in [5, 6]:  # Saturday or Sunday
            weekend_count += 1  # number of weekends in month
            ws[f"A{i}"].font = Font(color="FF0000")  # Red font for "Start"
            ws[f"B{i}"].font = Font(color="FF0000")  # Red font for "End"
            ws[f"C{i}"] = ""  # Empty values for weekends
            ws[f"D{i}"] = ""  # Empty values for weekends
            ws[f"E{i}"] = ""  # Empty values for weekends
            ws[f"F{i}"] = ""  # Empty values for weekends
            ws[f"G{i}"] = ""  # Empty values for weekends
            ws[f"H{i}"] = ""  # Empty values for weekends
            ws[f"I{i}"] = ""  # Empty values for weekends
            ws[f"J{i}"] = ""  # Empty values for weekends
            ws[f"K{i}"] = ""  # Empty values for weekends

        # Apply borders to each cell
        ws[f"A{i}"].border = thin_border_1
        ws[f"B{i}"].border = thin_border_1
        ws[f"C{i}"].border = thin_border_1
        ws[f"D{i}"].border = thin_border_1
        ws[f"E{i}"].border = thin_border_1
        ws[f"F{i}"].border = thin_border_1
        ws[f"G{i}"].border = thin_border_1
        ws[f"H{i}"].border = thin_border_1
        ws[f"I{i}"].border = thin_border_1
        ws[f"J{i}"].border = thin_border_1
        ws[f"K{i}"].border = thin_border_1

    # Set odometer reading as on last day every month
    ws["H8"] = last_odometer_value

    # Set total travel for the month
    ws["I7"] = ws["H8"].value - ws["H7"].value

    # Add 3 empty rows at the end
    for j in range(1, 5):  # Create 5 empty rows
        for col in range(1, 12):  # Assuming 11 columns (A to K)
            cell = ws.cell(row=i + j, column=col)  # Get the cell
            cell.value = ""  # Set the cell value to empty
            cell.border = thin_border_1  # Apply thin border to the cell

    # Get total spent for the month
    total_cost += WORK_RELATED_KM * INR_PER_KM * (int(last_day.strftime("%d")) - weekend_count)
    ws[f"K{i+j}"] = total_cost


    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # fix width for some columns
    ws.column_dimensions["A"].width = 20  # Set column A width to 20
    ws.column_dimensions["G"].width = 20  # Set column G width to 20
    ws.column_dimensions["H"].width = 20  # Set column H width to 20
    ws.column_dimensions["I"].width = 15  # Set column H width to 20


# Create sheets for each month
current_odometer = INITIAL_ODOMETER
for month in range(START_MONTH, 13):  # Apr-Dec 2024
    ws = workbook.create_sheet(title=datetime(2024, month, 1).strftime("%b%y"))
    create_month_sheet(ws, 2024, month, current_odometer)
    current_odometer += WORK_RELATED_KM * (pd.date_range(start=datetime(2024, month, 1),
                                                         end=(datetime(2024, month, 1) + timedelta(days=32)).replace(
                                                             day=1) - timedelta(days=1)).size)

for month in range(1, 4):  # Jan-Mar 2025
    ws = workbook.create_sheet(title=datetime(2025, month, 1).strftime("%b%y"))
    create_month_sheet(ws, 2025, month, current_odometer)
    current_odometer += WORK_RELATED_KM * (pd.date_range(start=datetime(2025, month, 1),
                                                         end=(datetime(2025, month, 1) + timedelta(days=32)).replace(
                                                             day=1) - timedelta(days=1)).size)

# Save the workbook
workbook.save(output_file_path)

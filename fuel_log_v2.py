import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import pandas as pd
import os
import argparse
import json
import logging

class FuelLogGenerator:
    """Class to generate fuel log workbooks for expense tracking"""

    def __init__(self, config=None):
        """
        Initialize with configuration settings
        
        Args:
            config (dict): Configuration dictionary with settings
        """
        # Setup logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.StreamHandler(),
                logging.FileHandler('fuel_log_generator.log')
            ]
        )
        self.logger = logging.getLogger('FuelLogGenerator')
        
        # Load default config
        self.config = {
            "start_date": datetime(2024, 8, 1),
            "end_date": datetime(2025, 3, 31),
            "initial_odometer": 17569,
            "inr_per_km": 10,
            "work_related_km": 110,
            "trip_purpose": "Official",
            "client_name": "Blink Charging",
            "is_work_travel": "Y",
            "personal_travel": "",
            "holidays": ['2025-01-26', '2025-03-10'],
            "employee": {
                "name": "Ashish Kumar",
                "id": "BLINKIN065",
                "department": "Technology",
                "manager": "Ajay Singh"
            },
            "vehicle": {
                "make": "Hyundai",
                "model": "Xcent",
                "year": "2018",
                "registration": "Delhi",
                "engine_size": "1199 CC"
            },
            "output_file_path": "Financial_Year_2024_25_Log_Book.xlsx"
        }
        
        # Override with provided config
        if config:
            self._update_config(config)
            
        # Initialize workbook
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # Remove the default sheet
        
        # Define styles
        self._define_styles()
        
        # Process holidays into datetime objects
        self._process_holidays()
        
    def _update_config(self, config):
        """Update configuration with provided values"""
        def deep_update(source, updates):
            for key, value in updates.items():
                if isinstance(value, dict) and key in source and isinstance(source[key], dict):
                    deep_update(source[key], value)
                else:
                    source[key] = value
        
        deep_update(self.config, config)
        
        # Convert string dates to datetime if needed
        if isinstance(self.config["start_date"], str):
            self.config["start_date"] = datetime.strptime(self.config["start_date"], "%Y-%m-%d")
        if isinstance(self.config["end_date"], str):
            self.config["end_date"] = datetime.strptime(self.config["end_date"], "%Y-%m-%d")
    
    def _process_holidays(self):
        """Process holidays from strings to datetime objects"""
        processed_holidays = []
        for holiday_str in self.config["holidays"]:
            try:
                holiday_date = datetime.strptime(holiday_str, "%Y-%m-%d")
                processed_holidays.append(holiday_date)
                self.logger.info(f"Added holiday: {holiday_date.strftime('%Y-%m-%d')}")
            except ValueError:
                self.logger.error(f"Invalid holiday date format: {holiday_str}")
        
        self.config["holidays"] = processed_holidays
    
    def _define_styles(self):
        """Define common styles used in the workbook"""
        # Font styles
        self.font_title = Font(size=18, bold=True, name="Algerian")
        self.font_heading = Font(size=12, bold=True, underline="single")
        self.font_subheading = Font(size=11, bold=True)
        self.font_weekend = Font(color="FF0000")
        self.font_holiday = Font(color="0000FF", bold=True)  # Blue and bold for holidays
        
        # Fill styles
        self.fill_header = PatternFill(fill_type="solid", start_color="B4C7E7", end_color="B4C7E7")
        self.fill_holiday = PatternFill(fill_type="solid", start_color="E6E6FF", end_color="E6E6FF")  # Light blue for holidays
        
        # Border styles
        self.border_all = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        self.border_right = Border(right=Side(border_style="thin", color="000000"))
        
        # Alignment styles
        self.align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        self.align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _apply_cell_style(self, ws, cell_ref, value, style_type):
        """Apply pre-defined styles to cells"""
        cell = ws[cell_ref]
        cell.value = value
        
        if style_type == "title":
            cell.font = self.font_title
            cell.fill = self.fill_header
            cell.border = self.border_all
            cell.alignment = self.align_center
        elif style_type == "section_heading":
            cell.font = self.font_heading
            cell.alignment = self.align_center
        elif style_type == "label":
            cell.font = self.font_subheading
            cell.alignment = self.align_right
        elif style_type == "value":
            cell.alignment = self.align_left
            cell.border = self.border_all
        elif style_type == "spacer":
            cell.border = self.border_right
        elif style_type == "header":
            cell.font = self.font_subheading
            cell.fill = self.fill_header
            cell.border = self.border_all
            cell.alignment = self.align_center
        
    def _create_month_sheet(self, year, month):
        """Create a worksheet for a given month"""
        month_date = datetime(year, month, 1)
        sheet_name = month_date.strftime("%b%y")
        self.logger.info(f"Creating sheet for {sheet_name}")
        
        ws = self.workbook.create_sheet(title=sheet_name)
        
        # Generate the date range for the month
        first_day = month_date
        last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        date_range = pd.date_range(start=first_day, end=last_day)
        
        # Calculate starting odometer for this month
        current_month_index = (year - self.config["start_date"].year) * 12 + (month - self.config["start_date"].month)
        
        if current_month_index == 0:
            odometer_start = self.config["initial_odometer"]
        else:
            # Calculate based on previous months
            odometer_start = self.config["initial_odometer"]
            for m_idx in range(current_month_index):
                m_year = self.config["start_date"].year + (self.config["start_date"].month + m_idx - 1) // 12
                m_month = (self.config["start_date"].month + m_idx - 1) % 12 + 1
                m_first_day = datetime(m_year, m_month, 1)
                m_last_day = (m_first_day + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                m_date_range = pd.date_range(start=m_first_day, end=m_last_day)
                
                # Count only workdays (not weekends or holidays)
                workday_count = 0
                for d in m_date_range:
                    if d.weekday() < 5 and not any(d.date() == h.date() for h in self.config["holidays"]):
                        workday_count += 1
                
                odometer_start += workday_count * self.config["work_related_km"]
        
        # Add headers and basic info
        self._add_sheet_headers(ws, year, month, odometer_start)
        
        # Fill in the data
        self._add_sheet_data(ws, date_range, odometer_start)
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
        
        return ws
    
    def _add_sheet_headers(self, ws, year, month, odometer_start):
        """Add headers and basic info to the worksheet"""
        # Company header
        self._apply_cell_style(ws, "A1", "Blink Charging Software Solutions India Private Limited", "title")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=11)
        
        # Basic details section
        self._apply_cell_style(ws, "A3", "BASIC DETAILS", "section_heading")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
        
        self._apply_cell_style(ws, "A5", "Name:", "label")
        self._apply_cell_style(ws, "B5", self.config["employee"]["name"], "value")
        self._apply_cell_style(ws, "A6", "Employee ID:", "label")
        self._apply_cell_style(ws, "B6", self.config["employee"]["id"], "value")
        self._apply_cell_style(ws, "A7", "Department:", "label")
        self._apply_cell_style(ws, "B7", self.config["employee"]["department"], "value")
        self._apply_cell_style(ws, "A8", "Manager:", "label")
        self._apply_cell_style(ws, "B8", self.config["employee"]["manager"], "value")
        
        # Vehicle details section
        self._apply_cell_style(ws, "D3", "VEHICLE DETAILS", "section_heading")
        ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=5)
        
        self._apply_cell_style(ws, "D4", "Make:", "label")
        self._apply_cell_style(ws, "E4", self.config["vehicle"]["make"], "value")
        self._apply_cell_style(ws, "D5", "Model:", "label")
        self._apply_cell_style(ws, "E5", self.config["vehicle"]["model"], "value")
        self._apply_cell_style(ws, "D6", "Year:", "label")
        self._apply_cell_style(ws, "E6", self.config["vehicle"]["year"], "value")
        self._apply_cell_style(ws, "D7", "Registration:", "label")
        self._apply_cell_style(ws, "E7", self.config["vehicle"]["registration"], "value")
        self._apply_cell_style(ws, "D8", "Engine Size:", "label")
        self._apply_cell_style(ws, "E8", self.config["vehicle"]["engine_size"], "value")
        
        # Odometer reading section
        self._apply_cell_style(ws, "G3", "ODOMETER READING", "section_heading")
        ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=8)
        
        self._apply_cell_style(ws, "G6", "Financial year:", "label")
        self._apply_cell_style(ws, "H6", f"{year}-{str(year+1)[-2:]}", "value")
        
        # Month details for odometer
        month_name = datetime(year, month, 1).strftime("%B")
        last_date = datetime(year, month, 1).replace(day=28) + timedelta(days=4)
        last_date = (last_date - timedelta(days=last_date.day)).day
        
        last_date_suffix = "th"
        if last_date == 1 or last_date == 21 or last_date == 31:
            last_date_suffix = "st"
        elif last_date == 2 or last_date == 22:
            last_date_suffix = "nd"
        elif last_date == 3 or last_date == 23:
            last_date_suffix = "rd"
            
        self._apply_cell_style(ws, "G7", f"As at 1st of {month_name} {year}", "label")
        self._apply_cell_style(ws, "G8", f"As at {last_date}{last_date_suffix} of {month_name} {year}", "label")
        
        self._apply_cell_style(ws, "H7", odometer_start, "value")
        # H8 will be filled after data is populated
        
        # Empty spacer columns
        for i in range(3, 10):
            self._apply_cell_style(ws, f"K{i}", "", "spacer")
        
        # Data table headers
        headers = [
            ("A10", "Date of Trip (DD/MM/YY)", "header"),
            ("C10", "Odometer Reading", "header"),
            ("E10", "Purpose of Trip", "header"),
            ("F10", "Name of Client", "header"),
            ("G10", "Work-related travel? (Y/N)", "header"),
            ("H10", "Work-related Travel (KM)", "header"),
            ("I10", "Personal Travel (KM)", "header"),
            ("J10", "INR Per KM", "header"),
            ("K10", "Amount (INR)", "header"),
            ("A11", "Start", "header"),
            ("B11", "End", "header"),
            ("C11", "Start", "header"),
            ("D11", "End", "header")
        ]
        
        for cell, value, style in headers:
            self._apply_cell_style(ws, cell, value, style)
        
        # Merge header cells
        ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=2)  # Date of Trip
        ws.merge_cells(start_row=10, start_column=3, end_row=10, end_column=4)  # Odometer Reading
        ws.merge_cells(start_row=10, start_column=5, end_row=11, end_column=5)  # Purpose of Trip
        ws.merge_cells(start_row=10, start_column=6, end_row=11, end_column=6)  # Name of Client
        ws.merge_cells(start_row=10, start_column=7, end_row=11, end_column=7)  # Work-related travel? (Y/N)
        ws.merge_cells(start_row=10, start_column=8, end_row=11, end_column=8)  # Work-related Travel (KM)
        ws.merge_cells(start_row=10, start_column=9, end_row=11, end_column=9)  # Personal Travel (KM)
        ws.merge_cells(start_row=10, start_column=10, end_row=11, end_column=10)  # INR Per KM
        ws.merge_cells(start_row=10, start_column=11, end_row=11, end_column=11)  # Amount (INR)
    
    def _add_sheet_data(self, ws, date_range, odometer_start):
        """Fill data rows for the given date range"""
        current_odometer = odometer_start
        total_cost = 0
        weekend_count = 0
        last_odometer_value = 0
        
        for i, date in enumerate(date_range, start=12):  # Start filling rows from the 12th row
            # Format the date
            date_str = date.strftime("%d/%m/%y")
            
            # Check if it's a weekend or holiday
            is_weekend = date.weekday() in [5, 6]  # Saturday or Sunday
            is_holiday = any(date.date() == h.date() for h in self.config["holidays"])
            
            # Apply appropriate styles
            for col in "ABCDEFGHIJK":
                cell = ws[f"{col}{i}"]
                cell.border = self.border_all
            
            if is_weekend or is_holiday:
                if is_weekend:
                    weekend_count += 1
                    ws[f"A{i}"].font = self.font_weekend
                    ws[f"B{i}"].font = self.font_weekend
                if is_holiday:
                    ws[f"A{i}"].font = self.font_holiday
                    ws[f"B{i}"].font = self.font_holiday
                    ws[f"A{i}"].fill = self.fill_holiday
                    ws[f"B{i}"].fill = self.fill_holiday
                
                ws[f"A{i}"].value = date_str
                ws[f"B{i}"].value = date_str
                
                # Empty cells for weekends and holidays
                for col in "CDEFGHIJK":
                    ws[f"{col}{i}"].value = ""
            else:
                # Add dates to the columns
                ws[f"A{i}"].value = date_str
                ws[f"B{i}"].value = date_str
                
                # Add odometer readings
                ws[f"C{i}"].value = current_odometer
                current_odometer += self.config["work_related_km"]
                ws[f"D{i}"].value = current_odometer
                
                # Add work-related travel details
                ws[f"E{i}"].value = self.config["trip_purpose"]
                ws[f"F{i}"].value = self.config["client_name"]
                ws[f"G{i}"].value = self.config["is_work_travel"]
                ws[f"H{i}"].value = self.config["work_related_km"]
                ws[f"I{i}"].value = self.config["personal_travel"]
                ws[f"J{i}"].value = self.config["inr_per_km"]
                ws[f"K{i}"].value = self.config["work_related_km"] * self.config["inr_per_km"]
                
                # Update total cost
                total_cost += self.config["work_related_km"] * self.config["inr_per_km"]
                
                # Save the last odometer value
                last_odometer_value = current_odometer
        
        # Add empty rows at the end
        for j in range(1, 5):  # Create 4 empty rows
            for col in "ABCDEFGHIJK":
                cell = ws.cell(row=i + j, column=ord(col) - 64)  # Convert letter to column number
                cell.value = ""
                cell.border = self.border_all
        
        # Add total for the month
        ws.cell(row=i + j, column=11).value = total_cost
        
        # Set the ending odometer reading
        ws["H8"] = last_odometer_value
        
        # Set total travel for the month
        ws["I7"] = last_odometer_value - odometer_start
    
    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths"""
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # Set a minimum width
            ws.column_dimensions[col_letter].width = max(max_length + 2, 12)
        
        # Fix width for specific columns
        ws.column_dimensions["A"].width = 20  # Date Start
        ws.column_dimensions["G"].width = 20  # Work-related travel
        ws.column_dimensions["H"].width = 20  # Work-related KM
        ws.column_dimensions["I"].width = 15  # Personal Travel
    
    def generate_workbook(self):
        """Generate the complete workbook with sheets for each month"""
        self.logger.info("Starting workbook generation")
        
        # Determine date range from config
        start_date = self.config["start_date"]
        end_date = self.config["end_date"]
        
        # Generate sheets for each month in the range
        current_date = start_date
        while current_date <= end_date:
            self._create_month_sheet(current_date.year, current_date.month)
            
            # Move to next month
            if current_date.month == 12:
                current_date = datetime(current_date.year + 1, 1, 1)
            else:
                current_date = datetime(current_date.year, current_date.month + 1, 1)
        
        # Save the workbook
        self.workbook.save(self.config["output_file_path"])
        self.logger.info(f"Workbook saved to {self.config['output_file_path']}")
        
    @classmethod
    def from_json_file(cls, json_file_path):
        """Create a FuelLogGenerator instance from a JSON configuration file"""
        try:
            with open(json_file_path, 'r') as file:
                config = json.load(file)
                return cls(config)
        except Exception as e:
            logging.error(f"Error loading configuration from {json_file_path}: {e}")
            return cls()


def main():
    """Main function to run the generator from command line"""
    parser = argparse.ArgumentParser(description='Generate a fuel log workbook for expense tracking.')
    parser.add_argument('--config', '-c', help='Path to JSON configuration file')
    parser.add_argument('--output', '-o', help='Output Excel file path')
    parser.add_argument('--start-date', help='Start date in YYYY-MM-DD format')
    parser.add_argument('--end-date', help='End date in YYYY-MM-DD format')
    parser.add_argument('--initial-odometer', type=int, help='Initial odometer reading')
    parser.add_argument('--km-per-day', type=int, help='Work-related kilometers per day')
    parser.add_argument('--rate-per-km', type=int, help='Rate per kilometer in INR')
    
    args = parser.parse_args()
    
    # Create generator with config file if provided
    if args.config:
        generator = FuelLogGenerator.from_json_file(args.config)
    else:
        generator = FuelLogGenerator()
    
    # Override with command line arguments
    config_overrides = {}
    
    if args.output:
        config_overrides["output_file_path"] = args.output
    if args.start_date:
        config_overrides["start_date"] = datetime.strptime(args.start_date, "%Y-%m-%d")
    if args.end_date:
        config_overrides["end_date"] = datetime.strptime(args.end_date, "%Y-%m-%d")
    if args.initial_odometer:
        config_overrides["initial_odometer"] = args.initial_odometer
    if args.km_per_day:
        config_overrides["work_related_km"] = args.km_per_day
    if args.rate_per_km:
        config_overrides["inr_per_km"] = args.rate_per_km
    
    if config_overrides:
        generator._update_config(config_overrides)
    
    # Generate the workbook
    generator.generate_workbook()


if __name__ == "__main__":
    main()
